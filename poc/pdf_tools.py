"""
PDF Tools — processing logic that runs entirely in the browser via Pyodide.

Nothing here touches the network or a server; it is loaded into a Pyodide
Web Worker and operates on file bytes handed over from the browser.

POC scope: two representative tools that exercise the whole stack.
  * compress_pdf   -> real lossless compression with pypdf (+ optional image
                      recompression via Pillow). Replaces the old, broken
                      byte-slicing "compressor" that produced corrupt PDFs.
  * table_to_pdf   -> Excel (.xlsx via openpyxl) or CSV (stdlib) -> PDF (fpdf2).
"""

import io
import csv

from pypdf import PdfReader, PdfWriter


def compress_pdf(data: bytes, image_quality: int = 60) -> bytes:
    """Losslessly compress content streams, drop duplicate objects, and
    recompress embedded raster images. Returns valid PDF bytes (never a
    truncated/corrupt file)."""
    reader = PdfReader(io.BytesIO(data))
    writer = PdfWriter()
    writer.append(reader)

    # 1. Lossless: deflate every page's content stream.
    for page in writer.pages:
        try:
            page.compress_content_streams()
        except Exception:
            pass  # some pages have nothing to compress

    # 2. Lossy-ish (opt-in): recompress embedded images via Pillow.
    if image_quality and image_quality < 100:
        for page in writer.pages:
            try:
                for img in page.images:
                    img.replace(img.image, quality=image_quality)
            except Exception:
                pass  # Pillow missing, or unsupported image filter

    # 3. Collapse identical/orphan objects where the pypdf version supports it.
    try:
        writer.compress_identical_objects()  # defaults dedupe + drop orphans
    except Exception:
        pass

    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def _rows_from_xlsx(data: bytes):
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c) for c in row])
    return rows


def _rows_from_csv(data: bytes):
    text = data.decode("utf-8", errors="replace")
    return [row for row in csv.reader(io.StringIO(text))]


def _latin1(s: str) -> str:
    # fpdf2 core fonts are latin-1 only; sanitize until we bundle a Unicode TTF.
    return str(s).encode("latin-1", "replace").decode("latin-1")


def table_to_pdf(data: bytes, is_csv: bool) -> bytes:
    from fpdf import FPDF

    rows = _rows_from_csv(data) if is_csv else _rows_from_xlsx(data)
    rows = [r for r in rows if any(str(c).strip() for c in r)] or [["(empty file)"]]

    ncols = max(len(r) for r in rows)
    rows = [[_latin1(c) for c in r] + [""] * (ncols - len(r)) for r in rows]

    pdf = FPDF(orientation="L", format="A4")
    pdf.set_auto_page_break(True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", size=8)
    with pdf.table(first_row_as_headings=True) as table:
        for r in rows:
            trow = table.row()
            for cell in r:
                trow.cell(cell)

    return bytes(pdf.output())


def run(path_in: str, path_out: str, action: str, is_csv: bool = False) -> int:
    """FS-based entry point used by the worker: read input file from Pyodide's
    virtual FS, process, write result. Returns output byte length."""
    with open(path_in, "rb") as f:
        data = f.read()

    if action == "compress":
        out = compress_pdf(data)
    elif action == "table":
        out = table_to_pdf(data, is_csv)
    else:
        raise ValueError(f"unknown action: {action}")

    with open(path_out, "wb") as f:
        f.write(out)
    return len(out)
